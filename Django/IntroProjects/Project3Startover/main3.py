#MODULES for directories, spreadsheets,
# machine learning and more
def paths():
    from pathlib import Path

    #Absolute path

    #relative path - path starting from the current directory
    #path = Path('Project4Ok')
    #print(path.rmdir()) #mkdir to make folder/directory

    path = Path()
    # print(path.glob('*.*')) #only get files but not directories
    # print(path.glob('*')) #to iterate over everthing
    #print(path.glob('*.py')) #search all py files

    #all above ^^👆👆👆 will bring generator object
    # to iterate over it use for loop eg.
    for file in path.glob('*'):
        print(file)


# WORKING WITH SPREADSHEETS
#wb as workbook
import openpyxl as xl
from openpyxl.chart import BarChart, Reference

#demo
def process_workbook():
    wb = xl.load_workbook('transactions.xlsx')
    sheet = wb['Sheet1'] #Sheet is case sensitive(capital)
    # cell = sheet['a1'] #one way to access a cell
    # cell = sheet.cell(1,1) #another way to access a cell
    # print(cell.value) #prints the cell stated above
    # print(sheet.max_row) #prints the no of rows in it

    for row in range(2, sheet.max_row + 1):
        cell = sheet.cell(row,3)
        corrected_price = cell.value * 0.9
        corrected_price_cell = sheet.cell(row,4)
        corrected_price_cell.value = corrected_price

    values = Reference(sheet, min_row = 2, max_row = sheet.max_row, min_col = 4,max_col = 4)

    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart, 'e2')


    wb.save('transactions2.xlsx')

#actual fuction to process many spreadsheets
def process_workbook(filename):
    
    wb = xl.load_workbook(filename)
    sheet = wb['Sheet1']
    
    for row in range(2, sheet.max_row + 1):
        cell = sheet.cell(row,3)
        corrected_price = cell.value * 0.9
        corrected_price_cell = sheet.cell(row,4)
        corrected_price_cell.value = corrected_price

    values = Reference(sheet, min_row = 2, max_row = sheet.max_row, min_col = 4,max_col = 4)

    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart, 'e2')


    wb.save(filename)




#mosh - machine learning(4:15:52)